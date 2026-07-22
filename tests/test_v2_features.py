from __future__ import annotations

import threading
import unittest
from dataclasses import replace
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from merge_engine import (
    MergeCancelled,
    MergeConfig,
    MergeError,
    SourceOverride,
    inspect_sources,
    merge_workbooks,
)
from merge_profiles import (
    CsvOptions,
    LabelRule,
    MergeProfile,
    OutputColumnRule,
    SpecialFieldRule,
)


def create_excel(
    path: Path,
    headers: list[object],
    rows: list[list[object]],
    *,
    sheet_name: str = "数据",
    leading_rows: int = 0,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for index in range(leading_rows):
        sheet.append([f"说明{index + 1}"])
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def simple_profile(
    *,
    optional_name: bool = False,
    label: LabelRule | None = None,
    specials: tuple[SpecialFieldRule, ...] = (),
    recognize: bool = False,
    recursive: bool = False,
) -> MergeProfile:
    return MergeProfile(
        name="自定义方案",
        output_columns=(
            OutputColumnRule("编号", aliases=("ID",), required=True),
            OutputColumnRule("姓名", aliases=("学员",), required=not optional_name),
        ),
        special_fields=specials,
        recognize_special_fields=recognize,
        label=label or LabelRule(False),
        recursive=recursive,
    ).validated()


def read_output_rows(path: Path) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    rows = list(workbook["合并结果"].iter_rows(values_only=True))
    workbook.close()
    return rows


class V2FeatureTests(unittest.TestCase):
    def test_custom_mapping_reorders_columns(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            create_excel(source / "数据.xlsx", ["姓名", "编号"], [["张三", 7]])
            profile = simple_profile()
            summary = merge_workbooks(MergeConfig(source, root / "out", "结果.xlsx", profile=profile))
            rows = read_output_rows(summary.output_files[0])
            self.assertEqual(rows, [("编号", "姓名"), (7, "张三")])

    def test_aliases_match_different_source_headers(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            create_excel(source / "数据.xlsx", ["ID", "学员"], [["001", "张三"]])
            summary = merge_workbooks(
                MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile())
            )
            self.assertEqual(read_output_rows(summary.output_files[0])[1], ("001", "张三"))

    def test_optional_missing_field_is_blank(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            create_excel(source / "数据.xlsx", ["编号"], [[1]])
            summary = merge_workbooks(
                MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile(optional_name=True))
            )
            self.assertEqual(read_output_rows(summary.output_files[0])[1], (1, None))

    def test_required_missing_field_skips_file(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            create_excel(source / "数据.xlsx", ["编号"], [[1]])
            plan = inspect_sources(MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile()))
            self.assertEqual(plan.files[0].status, "跳过")
            self.assertIn("姓名", plan.files[0].reason)

    def test_duplicate_headers_are_rejected_for_custom_mapping(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            create_excel(source / "数据.xlsx", ["编号", " 编号 ", "姓名"], [[1, 2, "甲"]])
            plan = inspect_sources(MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile()))
            self.assertEqual(plan.files[0].status, "跳过")
            self.assertIn("重复字段", plan.files[0].reason)

    def test_one_output_rule_matching_two_source_aliases_is_rejected(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            create_excel(source / "数据.xlsx", ["编号", "ID", "姓名"], [[1, 2, "甲"]])
            plan = inspect_sources(MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile()))
            self.assertEqual(plan.files[0].status, "跳过")
            self.assertIn("多个源字段", plan.files[0].reason)

    def test_multiple_special_fields_render_one_label(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            create_excel(
                source / "课程.xlsx",
                ["编号", "姓名", "主讲老师", "部门"],
                [[1, "甲", "张老师", "培训部"], [2, "乙", "张老师", "培训部"]],
            )
            profile = simple_profile(
                specials=(SpecialFieldRule("主讲老师"), SpecialFieldRule("部门")),
                recognize=True,
                label=LabelRule(
                    True,
                    "来源",
                    "{文件名不含扩展名}-{特殊字段:主讲老师}-{特殊字段:部门}",
                ),
            )
            summary = merge_workbooks(MergeConfig(source, root / "out", "结果.xlsx", profile=profile))
            rows = read_output_rows(summary.output_files[0])
            self.assertEqual(rows[0], ("编号", "姓名", "来源"))
            self.assertEqual(rows[1][-1], "课程-张老师-培训部")

    def test_multiple_values_in_special_field_skip_file(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            create_excel(
                source / "课程.xlsx",
                ["编号", "姓名", "老师"],
                [[1, "甲", "张"], [2, "乙", "李"]],
            )
            profile = simple_profile(
                specials=(SpecialFieldRule("老师"),),
                recognize=True,
            )
            plan = inspect_sources(MergeConfig(source, root / "out", "结果.xlsx", profile=profile))
            self.assertEqual(plan.files[0].status, "跳过")
            self.assertIn("多个值", plan.files[0].reason)

    def test_blank_special_field_skips_file(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            create_excel(source / "课程.xlsx", ["编号", "姓名", "老师"], [[1, "甲", None]])
            profile = simple_profile(specials=(SpecialFieldRule("老师"),), recognize=True)
            plan = inspect_sources(MergeConfig(source, root / "out", "结果.xlsx", profile=profile))
            self.assertEqual(plan.files[0].status, "跳过")
            self.assertIn("没有非空值", plan.files[0].reason)

    def test_label_can_be_disabled_without_special_field(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            create_excel(source / "课程.xlsx", ["编号", "姓名"], [[1, "甲"]])
            summary = merge_workbooks(
                MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile())
            )
            self.assertEqual(len(read_output_rows(summary.output_files[0])[0]), 2)

    def test_utf8_csv_preserves_leading_zeros_as_text(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            (source / "数据.csv").write_text("编号,姓名\n00123,张三\n", encoding="utf-8")
            summary = merge_workbooks(
                MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile())
            )
            value = read_output_rows(summary.output_files[0])[1][0]
            self.assertEqual(value, "00123")

    def test_utf8_bom_csv_is_detected(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            (source / "数据.csv").write_text("编号,姓名\n1,甲\n", encoding="utf-8-sig")
            plan = inspect_sources(MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile()))
            self.assertEqual(plan.files[0].csv_encoding, "utf-8-sig")
            self.assertEqual(plan.files[0].data_rows, 1)

    def test_gb18030_semicolon_csv_is_detected(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            (source / "数据.csv").write_text("编号;姓名\n0001;中文\n", encoding="gb18030")
            plan = inspect_sources(MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile()))
            self.assertEqual(plan.files[0].csv_encoding, "gb18030")
            self.assertEqual(plan.files[0].csv_delimiter, ";")

    def test_csv_header_row_and_tab_override(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            path = source / "数据.csv"
            path.write_text("导出说明\n编号\t姓名\n01\t甲\n", encoding="utf-8")
            override = SourceOverride(path, header_row=2, csv_encoding="utf-8", csv_delimiter="\t")
            config = MergeConfig(
                source,
                root / "out",
                "结果.xlsx",
                profile=simple_profile(),
                source_overrides=(override,),
            )
            summary = merge_workbooks(config)
            self.assertEqual(read_output_rows(summary.output_files[0])[1], ("01", "甲"))

    def test_csv_quoted_newline_is_one_data_row(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            (source / "数据.csv").write_text('编号,姓名\n1,"张\n三"\n', encoding="utf-8")
            path = source / "数据.csv"
            plan = inspect_sources(
                MergeConfig(
                    source,
                    root / "out",
                    "结果.xlsx",
                    profile=simple_profile(),
                    source_overrides=(SourceOverride(path, csv_delimiter=","),),
                )
            )
            self.assertEqual(plan.total_rows, 1)

    def test_recursive_scan_uses_relative_path_in_label(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            nested = source / "一部"
            nested.mkdir(parents=True)
            create_excel(nested / "数据.xlsx", ["编号", "姓名"], [[1, "甲"]])
            profile = simple_profile(
                recursive=True,
                label=LabelRule(True, "路径", "{相对路径}"),
            )
            summary = merge_workbooks(MergeConfig(source, root / "out", "结果.xlsx", profile=profile))
            self.assertEqual(read_output_rows(summary.output_files[0])[1][-1], "一部/数据.xlsx")

    def test_user_can_exclude_a_file(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            one = source / "一.xlsx"
            two = source / "二.xlsx"
            create_excel(one, ["编号", "姓名"], [[1, "甲"]])
            create_excel(two, ["编号", "姓名"], [[2, "乙"]])
            config = MergeConfig(
                source,
                root / "out",
                "结果.xlsx",
                profile=simple_profile(),
                source_overrides=(SourceOverride(two, include=False),),
            )
            summary = merge_workbooks(config)
            self.assertEqual(summary.total_rows, 1)
            self.assertEqual({item.status for item in summary.file_results}, {"成功", "排除"})

    def test_worksheet_override_selects_requested_visible_sheet(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            path = source / "多表.xlsx"
            workbook = Workbook()
            workbook.active.title = "说明"
            workbook.active.append(["不是表头"])
            data = workbook.create_sheet("业务数据")
            data.append(["编号", "姓名"])
            data.append([1, "甲"])
            workbook.save(path)
            workbook.close()
            config = MergeConfig(
                source,
                root / "out",
                "结果.xlsx",
                profile=simple_profile(),
                source_overrides=(SourceOverride(path, worksheet="业务数据"),),
            )
            summary = merge_workbooks(config)
            self.assertEqual(summary.total_rows, 1)
            self.assertEqual(summary.file_results[0].worksheet, "业务数据")

    def test_xlsm_input_is_read_without_preserving_macros(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            create_excel(source / "数据.xlsm", ["编号", "姓名"], [[1, "甲"]])
            summary = merge_workbooks(
                MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile())
            )
            self.assertEqual(summary.total_rows, 1)
            self.assertEqual(summary.file_results[0].file_type, "XLSM")

    def test_uncalculated_formula_cache_becomes_blank(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            create_excel(source / "公式.xlsx", ["编号", "姓名"], [[1, "=1+1"]])
            summary = merge_workbooks(
                MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile(optional_name=True))
            )
            self.assertEqual(read_output_rows(summary.output_files[0])[1], (1, None))

    def test_csv_illegal_xml_character_is_removed_and_reported(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            (source / "数据.csv").write_text("编号,姓名\n1,张\x01三\n", encoding="utf-8")
            summary = merge_workbooks(
                MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile())
            )
            self.assertEqual(read_output_rows(summary.output_files[0])[1][1], "张三")
            self.assertEqual(summary.file_results[0].adjusted_cells, 1)

    def test_long_csv_cell_is_truncated_to_excel_limit(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            long_name = "甲" * 40_000
            (source / "数据.csv").write_text(f"编号,姓名\n1,{long_name}\n", encoding="utf-8")
            summary = merge_workbooks(
                MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile())
            )
            self.assertEqual(len(read_output_rows(summary.output_files[0])[1][1]), 32_767)
            self.assertEqual(summary.file_results[0].adjusted_cells, 1)

    def test_changed_file_after_inspection_is_skipped(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            path = source / "数据.xlsx"
            create_excel(path, ["编号", "姓名"], [[1, "甲"]])
            config = MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile())
            plan = inspect_sources(config)
            create_excel(path, ["编号", "姓名"], [[1, "甲"], [2, "乙"]])
            summary = merge_workbooks(config, plan=plan)
            self.assertFalse(summary.output_files)
            self.assertEqual(summary.file_results[0].status, "跳过")
            self.assertIn("发生变化", summary.file_results[0].reason)

    def test_plan_with_different_profile_is_rejected(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            create_excel(source / "数据.xlsx", ["编号", "姓名"], [[1, "甲"]])
            config = MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile())
            plan = inspect_sources(config)
            other = replace(simple_profile(), name="另一个方案").validated()
            with self.assertRaisesRegex(MergeError, "方案不一致"):
                merge_workbooks(replace(config, profile=other), plan=plan)

    def test_all_invalid_files_still_generate_excel_report(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            create_excel(source / "数据.xlsx", ["编号"], [[1]])
            summary = merge_workbooks(
                MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile())
            )
            self.assertFalse(summary.output_files)
            self.assertTrue(summary.report_path.exists())
            workbook = load_workbook(summary.report_path, read_only=True)
            self.assertIn("文件明细", workbook.sheetnames)
            workbook.close()

    def test_report_treats_formula_like_filename_as_text(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            (source / "=危险.csv").write_text("编号,姓名\n1,甲\n", encoding="utf-8")
            summary = merge_workbooks(
                MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile())
            )
            report = load_workbook(summary.report_path, data_only=False)
            cell = report["文件明细"].cell(row=2, column=2)
            self.assertEqual(cell.value, "=危险.csv")
            self.assertEqual(cell.data_type, "s")
            report.close()

    def test_custom_mapping_splits_at_configured_boundary(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            create_excel(source / "数据.xlsx", ["编号", "姓名"], [[i, f"用户{i}"] for i in range(5)])
            summary = merge_workbooks(
                MergeConfig(
                    source,
                    root / "out",
                    "结果.xlsx",
                    max_data_rows_per_file=2,
                    profile=simple_profile(),
                )
            )
            self.assertEqual(len(summary.output_files), 3)
            self.assertEqual(sum(len(read_output_rows(path)) - 1 for path in summary.output_files), 5)

    def test_inspection_honors_cancellation(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "input"
            source.mkdir()
            create_excel(source / "数据.xlsx", ["编号", "姓名"], [[1, "甲"]])
            cancel = threading.Event()
            cancel.set()
            with self.assertRaises(MergeCancelled):
                inspect_sources(
                    MergeConfig(source, root / "out", "结果.xlsx", profile=simple_profile()),
                    cancel_event=cancel,
                )


if __name__ == "__main__":
    unittest.main()
