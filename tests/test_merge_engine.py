from __future__ import annotations

import threading
import unittest
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from merge_engine import (
    MergeCancelled,
    MergeConfig,
    OutputExistsError,
    SOURCE_HEADER,
    find_input_files,
    find_output_conflicts,
    inspect_sources,
    merge_workbooks,
    normalize_output_name,
)


def base_headers() -> list[str]:
    return ["编号", "姓名", "主讲老师"] + [f"字段{index}" for index in range(4, 18)]


def make_row(identifier: int, name: str, teacher: str) -> list[object]:
    return [identifier, name, teacher] + [f"值{identifier}-{index}" for index in range(4, 18)]


def create_workbook(
    path: Path,
    *,
    headers: list[object] | None = None,
    rows: list[list[object]] | None = None,
    hidden_leading_sheet: bool = False,
    literal_equals_cell: tuple[int, int] | None = None,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    if hidden_leading_sheet:
        sheet.title = "隐藏说明"
        sheet.sheet_state = "hidden"
        sheet = workbook.create_sheet("数据")
    else:
        sheet.title = "数据"
    sheet.append(headers or base_headers())
    for row in rows or []:
        sheet.append(row)
    if literal_equals_cell is not None:
        row_number, column_number = literal_equals_cell
        cell = sheet.cell(row=row_number, column=column_number)
        cell.value = "=这是文本"
        cell.data_type = "s"
    workbook.save(path)
    workbook.close()


class MergeEngineTests(unittest.TestCase):
    def test_normalize_output_name(self) -> None:
        self.assertEqual(normalize_output_name("结果"), "结果.xlsx")
        self.assertEqual(normalize_output_name("结果.XLSX"), "结果.XLSX")
        with self.assertRaisesRegex(Exception, "不能包含路径"):
            normalize_output_name("子目录/结果.xlsx")
        with self.assertRaisesRegex(Exception, "不能包含"):
            normalize_output_name("结果?.xlsx")

    def test_find_input_files_filters_sorts_and_does_not_recurse(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            (root / "B.xlsx").touch()
            (root / "a.XLSX").touch()
            (root / "c.csv").touch()
            (root / "d.xlsm").touch()
            (root / "~$锁定.xlsx").touch()
            (root / "说明.txt").touch()
            (root / "sub").mkdir()
            (root / "sub" / "z.xlsx").touch()
            excluded = root / "B.xlsx"
            files = find_input_files(root, excluded_paths=[excluded])
            self.assertEqual([path.name for path in files], ["a.XLSX", "c.csv", "d.xlsm"])

    def test_find_input_files_recurses_and_ignores_reports(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            (root / "sub").mkdir()
            (root / "sub" / "数据.csv").write_text("a\n1\n", encoding="utf-8")
            (root / "合并报告_20260101.xlsx").touch()
            files = find_input_files(root, recursive=True)
            self.assertEqual([item.relative_to(root).as_posix() for item in files], ["sub/数据.csv"])

    def test_basic_merge_preserves_values_and_adds_source_column(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()
            row_a = make_row(1, "甲", "张老师")
            row_a[3] = datetime(2026, 7, 22, 9, 30)
            create_workbook(
                input_dir / "课程A.xlsx",
                rows=[row_a, [None] * 17],
                literal_equals_cell=(2, 5),
            )
            create_workbook(
                input_dir / "课程B.xlsx",
                rows=[make_row(2, "乙", "李老师")],
                hidden_leading_sheet=True,
            )
            summary = merge_workbooks(MergeConfig(input_dir, output_dir, "结果.xlsx"))
            self.assertEqual(summary.total_rows, 2)
            self.assertEqual([path.name for path in summary.output_files], ["结果.xlsx"])
            self.assertEqual(summary.report_path.suffix, ".xlsx")
            workbook = load_workbook(summary.output_files[0], data_only=False)
            sheet = workbook["合并结果"]
            self.assertEqual(sheet.max_column, 18)
            self.assertEqual(sheet.max_row, 3)
            self.assertEqual(sheet.cell(1, 18).value, SOURCE_HEADER)
            self.assertEqual(sheet.cell(2, 18).value, "课程A - 张老师")
            self.assertEqual(sheet.cell(3, 18).value, "课程B - 李老师")
            self.assertEqual(sheet.cell(2, 5).value, "=这是文本")
            self.assertEqual(sheet.cell(2, 5).data_type, "s")
            self.assertIsInstance(sheet.cell(2, 4).value, datetime)
            self.assertEqual(sheet.freeze_panes, "A2")
            self.assertEqual(sheet.auto_filter.ref, "A1:R3")
            self.assertTrue(sheet.cell(1, 1).font.bold)
            workbook.close()
            report = load_workbook(summary.report_path, read_only=True, data_only=True)
            self.assertEqual(report.sheetnames, ["任务汇总", "文件明细", "字段映射"])
            values = dict(report["任务汇总"].iter_rows(min_row=2, values_only=True))
            self.assertEqual(values["成功文件数"], 2)
            self.assertEqual(values["合并数据行"], 2)
            report.close()

    def test_header_whitespace_is_normalized_for_comparison(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()
            create_workbook(input_dir / "01.xlsx", rows=[make_row(1, "甲", "张老师")])
            headers = base_headers()
            headers[0] = " 编号 "
            headers[2] = " 主讲老师 "
            create_workbook(input_dir / "02.xlsx", headers=headers, rows=[make_row(2, "乙", "李老师")])
            summary = merge_workbooks(MergeConfig(input_dir, output_dir, "结果.xlsx"))
            self.assertEqual(summary.total_rows, 2)
            self.assertTrue(all(result.status == "成功" for result in summary.file_results))

    def test_invalid_files_are_skipped_and_reported(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()
            create_workbook(input_dir / "01_有效.xlsx", rows=[make_row(1, "甲", "张老师")])
            missing = base_headers()
            missing[2] = "授课人"
            create_workbook(input_dir / "02_缺老师.xlsx", headers=missing, rows=[make_row(2, "乙", "李老师")])
            mismatch = base_headers()
            mismatch[0] = "不同编号"
            create_workbook(input_dir / "03_表头不同.xlsx", headers=mismatch, rows=[make_row(3, "丙", "王老师")])
            (input_dir / "04_损坏.xlsx").write_bytes(b"not-an-xlsx")
            create_workbook(input_dir / "05_空数据.xlsx", rows=[])
            summary = merge_workbooks(MergeConfig(input_dir, output_dir, "结果.xlsx"))
            self.assertEqual(summary.total_rows, 1)
            statuses = {result.path.name: result for result in summary.file_results}
            self.assertEqual(statuses["01_有效.xlsx"].status, "成功")
            for name in ("02_缺老师.xlsx", "03_表头不同.xlsx", "04_损坏.xlsx", "05_空数据.xlsx"):
                self.assertEqual(statuses[name].status, "跳过")
                self.assertTrue(statuses[name].reason)
            report = load_workbook(summary.report_path, read_only=True, data_only=True)
            rows = list(report["文件明细"].iter_rows(min_row=2, values_only=True))
            self.assertEqual(sum(1 for row in rows if row[0] == "跳过"), 4)
            self.assertTrue(any("A:Q 表头" in str(row[-1]) for row in rows))
            report.close()

    def test_splitting_uses_excel_row_boundary_and_repeats_header(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()
            rows = [make_row(index, f"用户{index}", "张老师") for index in range(1, 6)]
            create_workbook(input_dir / "课程.xlsx", rows=rows)
            summary = merge_workbooks(
                MergeConfig(input_dir, output_dir, "结果.xlsx", max_data_rows_per_file=2)
            )
            self.assertEqual(
                [path.name for path in summary.output_files],
                ["结果_第1部分.xlsx", "结果_第2部分.xlsx", "结果_第3部分.xlsx"],
            )
            for path, row_count in zip(summary.output_files, [3, 3, 2], strict=True):
                workbook = load_workbook(path, read_only=True)
                rows_out = list(workbook["合并结果"].iter_rows(values_only=True))
                self.assertEqual(len(rows_out), row_count)
                self.assertEqual(rows_out[0][17], SOURCE_HEADER)
                workbook.close()

    def test_existing_outputs_require_confirmation_and_stale_parts_are_removed(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()
            output_dir.mkdir()
            create_workbook(input_dir / "课程.xlsx", rows=[make_row(1, "甲", "张老师")])
            exact = output_dir / "结果.xlsx"
            stale = output_dir / "结果_第9部分.xlsx"
            exact.write_text("old exact", encoding="utf-8")
            stale.write_text("old stale", encoding="utf-8")
            conflicts = find_output_conflicts(output_dir, "结果.xlsx")
            self.assertEqual({path.name for path in conflicts}, {exact.name, stale.name})
            with self.assertRaises(OutputExistsError):
                merge_workbooks(MergeConfig(input_dir, output_dir, "结果.xlsx"))
            self.assertEqual(exact.read_text(encoding="utf-8"), "old exact")
            summary = merge_workbooks(MergeConfig(input_dir, output_dir, "结果.xlsx", overwrite=True))
            self.assertTrue(exact.exists())
            self.assertFalse(stale.exists())
            self.assertEqual({path.name for path in summary.overwritten_files}, {exact.name, stale.name})

    def test_cancellation_cleans_temporary_outputs(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()
            create_workbook(input_dir / "课程.xlsx", rows=[make_row(1, "甲", "张老师")])
            cancellation = threading.Event()
            cancellation.set()
            with self.assertRaises(MergeCancelled):
                merge_workbooks(
                    MergeConfig(input_dir, output_dir, "结果.xlsx"),
                    cancel_event=cancellation,
                )
            self.assertFalse((output_dir / "结果.xlsx").exists())
            self.assertFalse(any(output_dir.glob(".excel_merge_*")))

    def test_inspection_can_be_reused_by_merge(self) -> None:
        with TemporaryDirectory() as temp:
            root = Path(temp)
            input_dir = root / "input"
            output_dir = root / "output"
            input_dir.mkdir()
            create_workbook(input_dir / "课程.xlsx", rows=[make_row(1, "甲", "张老师")])
            config = MergeConfig(input_dir, output_dir, "结果.xlsx")
            plan = inspect_sources(config)
            self.assertEqual(plan.total_rows, 1)
            summary = merge_workbooks(config, plan=plan)
            self.assertEqual(summary.total_rows, 1)


if __name__ == "__main__":
    unittest.main()
