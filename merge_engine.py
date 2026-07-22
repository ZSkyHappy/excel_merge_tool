from __future__ import annotations

import csv
import re
import shutil
import tempfile
import threading
import time
import unicodedata
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Iterable, Iterator, Literal, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from merge_profiles import (
    APP_NAME,
    APP_VERSION,
    CsvOptions,
    MergeProfile,
    OutputColumnRule,
    ProfileError,
    SpecialFieldRule,
    classic_profile,
    create_profile_from_headers,
    normalize_header,
    render_label,
)


INPUT_COLUMN_COUNT = 17
OUTPUT_COLUMN_COUNT = 18
MAX_DATA_ROWS_PER_FILE = 1_048_575
TEACHER_HEADER = "主讲老师"
SOURCE_HEADER = "所属表格+主讲老师"
OUTPUT_SHEET_NAME = "合并结果"
PROGRESS_INTERVAL_ROWS = 1_000
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}

_INVALID_OUTPUT_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
_ILLEGAL_XML_CHARS = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")
_PART_NAME_TEMPLATE = "{stem}_第{number}部分.xlsx"
_REPORT_PREFIX = "合并报告_"
_WINDOWS_RESERVED_NAMES = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{index}" for index in range(1, 10)),
    *(f"LPT{index}" for index in range(1, 10)),
}


class MergeError(RuntimeError):
    """Base error for merge operations."""


class MergeCancelled(MergeError):
    """Raised after a cooperative cancellation request."""


class OutputExistsError(MergeError):
    """Raised when output files already exist and overwrite is disabled."""

    def __init__(self, paths: Sequence[Path]):
        self.paths = tuple(paths)
        names = "、".join(path.name for path in self.paths)
        super().__init__(f"输出目录中已存在同名结果：{names}")


Status = Literal["成功", "跳过", "排除"]
InspectionStatus = Literal["有效", "跳过", "排除"]
Phase = Literal[
    "scan_start",
    "scan_file",
    "scan_rows",
    "scan_done",
    "merge_start",
    "merge_file",
    "merge_rows",
    "saving",
    "done",
]


@dataclass(frozen=True)
class SourceOverride:
    path: Path
    include: bool = True
    worksheet: str = ""
    header_row: int | None = None
    csv_encoding: str = ""
    csv_delimiter: str = ""

    def validated(self) -> "SourceOverride":
        header_row = self.header_row
        if header_row is not None and (int(header_row) < 1 or int(header_row) > 1_048_576):
            raise MergeError("逐文件表头行必须在 1 到 1,048,576 之间。")
        options = CsvOptions(
            encoding=self.csv_encoding or "auto",
            delimiter=self.csv_delimiter or "auto",
        ).validated()
        return SourceOverride(
            path=Path(self.path).expanduser().resolve(),
            include=bool(self.include),
            worksheet=str(self.worksheet or "").strip(),
            header_row=int(header_row) if header_row is not None else None,
            csv_encoding=options.encoding if self.csv_encoding else "",
            csv_delimiter=options.delimiter if self.csv_delimiter else "",
        )


@dataclass(frozen=True)
class MergeConfig:
    input_dir: Path
    output_dir: Path
    output_name: str
    overwrite: bool = False
    max_data_rows_per_file: int = MAX_DATA_ROWS_PER_FILE
    progress_interval_rows: int = PROGRESS_INTERVAL_ROWS
    profile: MergeProfile | None = None
    source_overrides: tuple[SourceOverride, ...] = field(default_factory=tuple)

    def validated(self) -> "MergeConfig":
        input_dir = Path(self.input_dir).expanduser().resolve()
        output_dir = Path(self.output_dir).expanduser().resolve()
        output_name = normalize_output_name(self.output_name)
        if not input_dir.is_dir():
            raise MergeError(f"输入文件夹不存在：{input_dir}")
        if self.max_data_rows_per_file < 1 or self.max_data_rows_per_file > MAX_DATA_ROWS_PER_FILE:
            raise MergeError(
                f"每个文件的数据行上限必须在 1 到 {MAX_DATA_ROWS_PER_FILE:,} 之间。"
            )
        if self.progress_interval_rows < 1:
            raise MergeError("进度更新间隔必须大于 0。")
        try:
            profile = (self.profile or classic_profile()).validated()
            overrides = tuple(item.validated() for item in self.source_overrides)
        except ProfileError as exc:
            raise MergeError(str(exc)) from exc
        return MergeConfig(
            input_dir=input_dir,
            output_dir=output_dir,
            output_name=output_name,
            overwrite=bool(self.overwrite),
            max_data_rows_per_file=int(self.max_data_rows_per_file),
            progress_interval_rows=int(self.progress_interval_rows),
            profile=profile,
            source_overrides=overrides,
        )


@dataclass(frozen=True)
class ProgressEvent:
    phase: Phase
    message: str
    current_file: str = ""
    files_completed: int = 0
    files_total: int = 0
    rows_written: int = 0
    total_rows: int = 0


@dataclass
class FileResult:
    path: Path
    status: Status
    rows_written: int = 0
    teacher: str = ""
    worksheet: str = ""
    reason: str = ""
    adjusted_cells: int = 0
    relative_path: str = ""
    file_type: str = ""
    header_row: int = 1
    csv_encoding: str = ""
    csv_delimiter: str = ""
    special_values: tuple[tuple[str, str], ...] = field(default_factory=tuple)


@dataclass(frozen=True)
class MergeSummary:
    output_files: tuple[Path, ...]
    report_path: Path
    file_results: tuple[FileResult, ...]
    total_rows: int
    elapsed_seconds: float
    overwritten_files: tuple[Path, ...] = field(default_factory=tuple)


@dataclass(frozen=True)
class SourceHeader:
    path: Path
    file_type: str
    headers: tuple[object, ...]
    worksheet: str = ""
    visible_sheets: tuple[str, ...] = field(default_factory=tuple)
    header_row: int = 1
    csv_encoding: str = ""
    csv_delimiter: str = ""


@dataclass(frozen=True)
class InspectedFile:
    path: Path
    relative_path: str
    file_type: str
    status: InspectionStatus
    reason: str = ""
    worksheet: str = ""
    visible_sheets: tuple[str, ...] = field(default_factory=tuple)
    header_row: int = 1
    csv_encoding: str = ""
    csv_delimiter: str = ""
    headers: tuple[object, ...] = field(default_factory=tuple)
    mapping: tuple[int | None, ...] = field(default_factory=tuple)
    special_mapping: tuple[int, ...] = field(default_factory=tuple)
    special_values: tuple[tuple[str, str], ...] = field(default_factory=tuple)
    data_rows: int = 0
    size: int = 0
    modified_ns: int = 0

    @property
    def special_values_dict(self) -> dict[str, str]:
        return dict(self.special_values)


@dataclass(frozen=True)
class InspectionPlan:
    input_dir: Path
    profile: MergeProfile
    files: tuple[InspectedFile, ...]
    output_headers: tuple[object, ...]
    column_widths: tuple[float, ...]
    total_rows: int
    created_at: datetime

    @property
    def valid_files(self) -> tuple[InspectedFile, ...]:
        return tuple(item for item in self.files if item.status == "有效")


ProgressCallback = Callable[[ProgressEvent], None]


def normalize_output_name(output_name: str) -> str:
    name = str(output_name).strip()
    if not name:
        raise MergeError("请输入输出文件名。")
    if name in {".", ".."} or Path(name).name != name:
        raise MergeError("输出文件名不能包含路径。")
    if _INVALID_OUTPUT_CHARS.search(name):
        raise MergeError('输出文件名不能包含 < > : " / \\ | ? * 等字符。')
    if not name.casefold().endswith(".xlsx"):
        name += ".xlsx"
    if not Path(name).stem.strip(" ."):
        raise MergeError("输出文件名无效。")
    if Path(name).stem.rstrip(" .").upper() in _WINDOWS_RESERVED_NAMES:
        raise MergeError("输出文件名是 Windows 保留名称，请更换名称。")
    return name


def find_input_files(
    input_dir: Path,
    *,
    excluded_paths: Iterable[Path] = (),
    recursive: bool = False,
) -> list[Path]:
    root = Path(input_dir).resolve()
    excluded = {Path(path).resolve() for path in excluded_paths}
    iterator = root.rglob("*") if recursive else root.iterdir()
    files = [
        path.resolve()
        for path in iterator
        if path.is_file()
        and path.suffix.casefold() in SUPPORTED_EXTENSIONS
        and not path.name.startswith("~$")
        and not (path.suffix.casefold() == ".xlsx" and path.name.startswith(_REPORT_PREFIX))
        and path.resolve() not in excluded
        and not any(part.startswith(".excel_merge_") for part in path.relative_to(root).parts)
    ]
    return sorted(
        files,
        key=lambda path: (
            path.relative_to(root).as_posix().casefold(),
            path.relative_to(root).as_posix(),
        ),
    )


def find_output_conflicts(output_dir: Path, output_name: str) -> list[Path]:
    directory = Path(output_dir).expanduser().resolve()
    if not directory.is_dir():
        return []
    normalized_name = normalize_output_name(output_name)
    stem = Path(normalized_name).stem
    exact_name = normalized_name.casefold()
    part_pattern = re.compile(
        rf"^{re.escape(stem)}_第\d+部分\.xlsx$",
        flags=re.IGNORECASE,
    )
    matches = [
        path.resolve()
        for path in directory.iterdir()
        if path.is_file()
        and (path.name.casefold() == exact_name or part_pattern.fullmatch(path.name))
    ]
    return sorted(matches, key=lambda path: (path.name.casefold(), path.name))


def list_visible_sheets(path: Path) -> tuple[str, ...]:
    source = Path(path)
    if source.suffix.casefold() not in {".xlsx", ".xlsm"}:
        return ()
    workbook = None
    try:
        workbook = load_workbook(source, read_only=True, data_only=True, keep_links=False)
        return tuple(sheet.title for sheet in workbook.worksheets if sheet.sheet_state == "visible")
    except Exception as exc:
        raise MergeError(f"无法读取工作表列表：{exc}") from exc
    finally:
        if workbook is not None:
            workbook.close()


def read_source_header(
    path: Path,
    *,
    override: SourceOverride | None = None,
    default_header_row: int = 1,
    csv_options: CsvOptions | None = None,
) -> SourceHeader:
    source = Path(path).expanduser().resolve()
    selected = (override or SourceOverride(source)).validated()
    header_row = selected.header_row or int(default_header_row)
    base_options = csv_options or CsvOptions()
    options = CsvOptions(
        encoding=selected.csv_encoding or base_options.encoding,
        delimiter=selected.csv_delimiter or base_options.delimiter,
    ).validated()
    reader = _open_source_reader(
        source,
        worksheet=selected.worksheet,
        header_row=header_row,
        csv_options=options,
    )
    try:
        return SourceHeader(
            path=source,
            file_type=reader.file_type,
            headers=reader.headers,
            worksheet=reader.worksheet,
            visible_sheets=reader.visible_sheets,
            header_row=reader.header_row,
            csv_encoding=reader.csv_encoding,
            csv_delimiter=reader.csv_delimiter,
        )
    finally:
        reader.close()


def create_profile_from_template(
    path: Path,
    name: str,
    *,
    override: SourceOverride | None = None,
    default_header_row: int = 1,
) -> tuple[MergeProfile, SourceHeader]:
    header = read_source_header(path, override=override, default_header_row=default_header_row)
    profile = create_profile_from_headers(name, header.headers, header_row=header.header_row)
    return profile, header


def inspect_sources(
    config: MergeConfig,
    *,
    cancel_event: threading.Event | None = None,
    progress_callback: ProgressCallback | None = None,
) -> InspectionPlan:
    cfg = config.validated()
    profile = cfg.profile or classic_profile()
    cancellation = cancel_event or threading.Event()
    conflicts = find_output_conflicts(cfg.output_dir, cfg.output_name)
    excluded = conflicts if cfg.input_dir == cfg.output_dir else []
    input_files = find_input_files(
        cfg.input_dir,
        excluded_paths=excluded,
        recursive=profile.recursive,
    )
    if not input_files:
        raise MergeError("所选文件夹中没有可处理的 .xlsx、.xlsm 或 .csv 文件。")

    override_map = {item.path.resolve(): item for item in cfg.source_overrides}
    _emit(
        progress_callback,
        ProgressEvent(
            phase="scan_start",
            message=f"开始精确预检 {len(input_files)} 个文件……",
            files_total=len(input_files),
        ),
    )
    inspected: list[InspectedFile] = []
    canonical_header: tuple[object, ...] | None = None
    canonical_normalized: tuple[str, ...] | None = None
    width_count = INPUT_COLUMN_COUNT if profile.classic else len(profile.output_columns)
    max_widths = [8.0] * (width_count + int(profile.label.enabled))

    for file_index, path in enumerate(input_files, start=1):
        _check_cancelled(cancellation)
        relative_path = path.relative_to(cfg.input_dir).as_posix()
        override = override_map.get(path.resolve(), SourceOverride(path)).validated()
        if not override.include:
            inspected.append(
                InspectedFile(
                    path=path,
                    relative_path=relative_path,
                    file_type=path.suffix.casefold().lstrip(".").upper(),
                    status="排除",
                    reason="用户已在预览中排除。",
                    header_row=override.header_row or profile.default_header_row,
                )
            )
            continue

        _emit(
            progress_callback,
            ProgressEvent(
                phase="scan_file",
                message=f"正在预检：{relative_path}",
                current_file=relative_path,
                files_completed=file_index - 1,
                files_total=len(input_files),
            ),
        )
        reader: _SourceReader | None = None
        try:
            stat_before = path.stat()
            options = CsvOptions(
                encoding=override.csv_encoding or profile.csv_options.encoding,
                delimiter=override.csv_delimiter or profile.csv_options.delimiter,
            ).validated()
            reader = _open_source_reader(
                path,
                worksheet=override.worksheet,
                header_row=override.header_row or profile.default_header_row,
                csv_options=options,
            )
            if not reader.headers or _is_blank_row(reader.headers):
                raise MergeError(f"第 {reader.header_row} 行表头为空。")

            if profile.classic:
                output_header = _pad_row(reader.headers, INPUT_COLUMN_COUNT)
                normalized_output = tuple(normalize_header(value) for value in output_header)
                mapping = tuple(range(INPUT_COLUMN_COUNT))
            else:
                header_index = _build_unique_header_index(reader.headers)
                mapping = _resolve_output_mapping(header_index, profile.output_columns)
                output_header = tuple(item.name for item in profile.output_columns)
                normalized_output = tuple(normalize_header(value) for value in output_header)

            special_mapping = (
                _resolve_special_mapping(
                    reader.headers,
                    profile.special_fields,
                    custom=not profile.classic,
                )
                if profile.recognize_special_fields
                else ()
            )
            max_col = max(
                [len(reader.headers), INPUT_COLUMN_COUNT if profile.classic else 0]
                + [index + 1 for index in mapping if index is not None]
                + [index + 1 for index in special_mapping]
            )
            data_rows = 0
            local_widths = [8.0] * width_count
            special_sets: list[set[str]] = [set() for _ in special_mapping]
            for row_number, row in enumerate(reader.iter_data_rows(max_col=max_col), start=reader.header_row + 1):
                if row_number % cfg.progress_interval_rows == 0:
                    _check_cancelled(cancellation)
                    _emit(
                        progress_callback,
                        ProgressEvent(
                            phase="scan_rows",
                            message=f"正在预检：{relative_path}（第 {row_number:,} 行）",
                            current_file=relative_path,
                            files_completed=file_index - 1,
                            files_total=len(input_files),
                        ),
                    )
                mapped = _map_row(row, mapping)
                if not _is_blank_row(mapped):
                    data_rows += 1
                    for index, value in enumerate(mapped):
                        local_widths[index] = max(local_widths[index], _display_width(value))
                for special_index, source_index in enumerate(special_mapping):
                    if source_index < len(row):
                        value = _value_as_text(row[source_index])
                        if value and len(special_sets[special_index]) < 2:
                            special_sets[special_index].add(value)

            if data_rows == 0:
                raise MergeError("映射后的数据区没有有效数据行。")
            special_values: list[tuple[str, str]] = []
            if profile.recognize_special_fields:
                for rule, values in zip(profile.special_fields, special_sets, strict=True):
                    if not values:
                        raise MergeError(f'特殊字段“{rule.name}”没有非空值。')
                    if len(values) > 1:
                        shown = "、".join(sorted(values))
                        raise MergeError(f'特殊字段“{rule.name}”存在多个值：{shown}')
                    special_values.append((rule.name, next(iter(values))))

            if profile.classic:
                if canonical_normalized is None:
                    canonical_header = tuple(output_header)
                    canonical_normalized = normalized_output
                elif normalized_output != canonical_normalized:
                    raise MergeError("A:Q 表头与首个有效文件不一致。")

            stat_after = path.stat()
            if stat_after.st_size != stat_before.st_size or stat_after.st_mtime_ns != stat_before.st_mtime_ns:
                raise MergeError("文件在预检过程中发生变化，请重新扫描。")
            for index, value in enumerate(output_header):
                if index < len(max_widths):
                    max_widths[index] = max(max_widths[index], _display_width(value))
            for index, width in enumerate(local_widths):
                max_widths[index] = max(max_widths[index], width)
            if profile.label.enabled:
                label_value = render_label(
                    profile.label,
                    file_name=path.name,
                    relative_path=relative_path,
                    worksheet=reader.worksheet,
                    special_values=dict(special_values),
                )
                max_widths[-1] = max(
                    max_widths[-1],
                    _display_width(profile.label.header),
                    _display_width(label_value),
                )
            inspected.append(
                InspectedFile(
                    path=path,
                    relative_path=relative_path,
                    file_type=reader.file_type,
                    status="有效",
                    worksheet=reader.worksheet,
                    visible_sheets=reader.visible_sheets,
                    header_row=reader.header_row,
                    csv_encoding=reader.csv_encoding,
                    csv_delimiter=reader.csv_delimiter,
                    headers=reader.headers,
                    mapping=mapping,
                    special_mapping=special_mapping,
                    special_values=tuple(special_values),
                    data_rows=data_rows,
                    size=stat_after.st_size,
                    modified_ns=stat_after.st_mtime_ns,
                )
            )
        except MergeCancelled:
            raise
        except Exception as exc:
            inspected.append(
                InspectedFile(
                    path=path,
                    relative_path=relative_path,
                    file_type=path.suffix.casefold().lstrip(".").upper(),
                    status="跳过",
                    reason=str(exc).strip() or exc.__class__.__name__,
                    worksheet=reader.worksheet if reader else override.worksheet,
                    visible_sheets=reader.visible_sheets if reader else (),
                    header_row=reader.header_row if reader else (override.header_row or profile.default_header_row),
                    csv_encoding=reader.csv_encoding if reader else override.csv_encoding,
                    csv_delimiter=reader.csv_delimiter if reader else override.csv_delimiter,
                )
            )
        finally:
            if reader is not None:
                reader.close()
        _emit(
            progress_callback,
            ProgressEvent(
                phase="scan_file",
                message=f"已预检 {file_index}/{len(input_files)} 个文件",
                current_file=relative_path,
                files_completed=file_index,
                files_total=len(input_files),
            ),
        )

    if profile.classic:
        data_headers = canonical_header or tuple(
            f"列{get_column_letter(index)}" for index in range(1, INPUT_COLUMN_COUNT + 1)
        )
    else:
        data_headers = tuple(item.name for item in profile.output_columns)
    output_headers = tuple(data_headers) + ((profile.label.header,) if profile.label.enabled else ())
    capped_widths = tuple(min(max(width + 2, 10), 42) for width in max_widths)
    total_rows = sum(item.data_rows for item in inspected if item.status == "有效")
    valid_count = sum(1 for item in inspected if item.status == "有效")
    skipped_count = sum(1 for item in inspected if item.status == "跳过")
    _emit(
        progress_callback,
        ProgressEvent(
            phase="scan_done",
            message=f"预检完成：{valid_count} 个有效，{skipped_count} 个跳过。",
            files_completed=len(input_files),
            files_total=len(input_files),
            total_rows=total_rows,
        ),
    )
    return InspectionPlan(
        input_dir=cfg.input_dir,
        profile=profile,
        files=tuple(inspected),
        output_headers=output_headers,
        column_widths=capped_widths,
        total_rows=total_rows,
        created_at=datetime.now(),
    )


def merge_workbooks(
    config: MergeConfig,
    *,
    plan: InspectionPlan | None = None,
    cancel_event: threading.Event | None = None,
    progress_callback: ProgressCallback | None = None,
) -> MergeSummary:
    started = time.perf_counter()
    cfg = config.validated()
    cfg.output_dir.mkdir(parents=True, exist_ok=True)
    cancellation = cancel_event or threading.Event()
    inspection = plan or inspect_sources(
        cfg,
        cancel_event=cancellation,
        progress_callback=progress_callback,
    )
    _validate_plan(cfg, inspection)
    conflicts = find_output_conflicts(cfg.output_dir, cfg.output_name)
    if conflicts and not cfg.overwrite:
        raise OutputExistsError(conflicts)

    temp_dir = Path(tempfile.mkdtemp(prefix=".excel_merge_", dir=str(cfg.output_dir))).resolve()
    writer: _OutputWriter | None = None
    published = False
    try:
        results = [_result_from_inspected(item) for item in inspection.files]
        result_map = {item.path: item for item in results}
        ready = list(inspection.valid_files)
        _emit(
            progress_callback,
            ProgressEvent(
                phase="merge_start",
                message=f"准备合并 {len(ready)} 个有效文件。",
                files_total=len(ready),
                total_rows=inspection.total_rows,
            ),
        )
        writer = _OutputWriter(
            temp_dir=temp_dir,
            headers=inspection.output_headers,
            column_widths=inspection.column_widths,
            max_data_rows=cfg.max_data_rows_per_file,
        )
        total_written = 0
        files_completed = 0
        for item in ready:
            _check_cancelled(cancellation)
            result = result_map[item.path]
            try:
                current_stat = item.path.stat()
            except OSError as exc:
                result.status = "跳过"
                result.reason = f"合并前无法读取文件状态：{exc}"
                continue
            if current_stat.st_size != item.size or current_stat.st_mtime_ns != item.modified_ns:
                result.status = "跳过"
                result.reason = "文件在预检后发生变化，请重新扫描。"
                continue

            _emit(
                progress_callback,
                ProgressEvent(
                    phase="merge_file",
                    message=f"正在合并：{item.relative_path}",
                    current_file=item.relative_path,
                    files_completed=files_completed,
                    files_total=len(ready),
                    rows_written=total_written,
                    total_rows=inspection.total_rows,
                ),
            )
            reader: _SourceReader | None = None
            file_written = 0
            adjusted_cells = 0
            try:
                reader = _open_source_reader(
                    item.path,
                    worksheet=item.worksheet,
                    header_row=item.header_row,
                    csv_options=CsvOptions(
                        encoding=item.csv_encoding or "auto",
                        delimiter=item.csv_delimiter or "auto",
                    ),
                )
                if tuple(normalize_header(value) for value in reader.headers) != tuple(
                    normalize_header(value) for value in item.headers
                ):
                    raise MergeError("文件表头与预检结果不同，请重新扫描。")
                label_value = ""
                if inspection.profile.label.enabled:
                    label_value = render_label(
                        inspection.profile.label,
                        file_name=item.path.name,
                        relative_path=item.relative_path,
                        worksheet=item.worksheet,
                        special_values=item.special_values_dict,
                    )
                max_col = max(
                    [len(item.headers)] + [index + 1 for index in item.mapping if index is not None]
                )
                for row_number, row in enumerate(
                    reader.iter_data_rows(max_col=max_col),
                    start=item.header_row + 1,
                ):
                    if row_number % cfg.progress_interval_rows == 0:
                        _check_cancelled(cancellation)
                    mapped = _map_row(row, item.mapping)
                    if _is_blank_row(mapped):
                        continue
                    prepared, changed = _prepare_output_values(mapped)
                    adjusted_cells += changed
                    output_values = prepared + ((label_value,) if inspection.profile.label.enabled else ())
                    writer.append(output_values)
                    file_written += 1
                    total_written += 1
                    if total_written % cfg.progress_interval_rows == 0:
                        _emit(
                            progress_callback,
                            ProgressEvent(
                                phase="merge_rows",
                                message=f"已写入 {total_written:,} 行",
                                current_file=item.relative_path,
                                files_completed=files_completed,
                                files_total=len(ready),
                                rows_written=total_written,
                                total_rows=inspection.total_rows,
                            ),
                        )
            except MergeCancelled:
                raise
            except Exception as exc:
                raise MergeError(
                    f"合并 {item.relative_path} 时发生错误。为避免输出部分数据，本批次已取消：{exc}"
                ) from exc
            finally:
                if reader is not None:
                    reader.close()
            result.rows_written = file_written
            result.adjusted_cells = adjusted_cells
            files_completed += 1
            _emit(
                progress_callback,
                ProgressEvent(
                    phase="merge_rows",
                    message=f"已完成：{item.relative_path}（{file_written:,} 行）",
                    current_file=item.relative_path,
                    files_completed=files_completed,
                    files_total=len(ready),
                    rows_written=total_written,
                    total_rows=inspection.total_rows,
                ),
            )

        _check_cancelled(cancellation)
        _emit(
            progress_callback,
            ProgressEvent(
                phase="saving",
                message="正在保存结果与 Excel 报告，请稍候……",
                files_completed=files_completed,
                files_total=len(ready),
                rows_written=total_written,
                total_rows=inspection.total_rows,
            ),
        )
        part_files = writer.finish()
        writer = None
        target_paths = _target_output_paths(cfg.output_dir, cfg.output_name, len(part_files))
        report_name = _unique_report_name(cfg.output_dir)
        report_temp_path = temp_dir / report_name
        report_target_path = cfg.output_dir / report_name
        elapsed = time.perf_counter() - started
        _write_report_workbook(
            report_temp_path,
            config=cfg,
            plan=inspection,
            results=results,
            output_paths=target_paths,
            total_rows=total_written,
            elapsed_seconds=elapsed,
            overwritten_files=conflicts if target_paths else (),
        )
        published_outputs, overwritten = _publish_results(
            part_files=part_files,
            target_paths=target_paths,
            report_temp_path=report_temp_path,
            report_target_path=report_target_path,
            conflicts=conflicts if target_paths else (),
            overwrite=cfg.overwrite,
            temp_dir=temp_dir,
        )
        published = True
        summary = MergeSummary(
            output_files=tuple(published_outputs),
            report_path=report_target_path,
            file_results=tuple(results),
            total_rows=total_written,
            elapsed_seconds=elapsed,
            overwritten_files=tuple(overwritten),
        )
        _emit(
            progress_callback,
            ProgressEvent(
                phase="done",
                message=f"合并完成，共写入 {total_written:,} 行。",
                files_completed=files_completed,
                files_total=len(ready),
                rows_written=total_written,
                total_rows=inspection.total_rows,
            ),
        )
        return summary
    finally:
        if writer is not None:
            writer.abort()
        if not published or temp_dir.exists():
            _safe_remove_temp_dir(temp_dir, cfg.output_dir)


def _validate_plan(config: MergeConfig, plan: InspectionPlan) -> None:
    if plan.input_dir.resolve() != config.input_dir.resolve():
        raise MergeError("预检计划与当前输入文件夹不一致，请重新扫描。")
    profile = config.profile or classic_profile()
    if plan.profile.to_dict() != profile.to_dict():
        raise MergeError("预检计划与当前合并方案不一致，请重新扫描。")


def _result_from_inspected(item: InspectedFile) -> FileResult:
    status: Status = "成功" if item.status == "有效" else ("排除" if item.status == "排除" else "跳过")
    teacher = dict(item.special_values).get(TEACHER_HEADER, "")
    return FileResult(
        path=item.path,
        status=status,
        teacher=teacher,
        worksheet=item.worksheet,
        reason=item.reason,
        relative_path=item.relative_path,
        file_type=item.file_type,
        header_row=item.header_row,
        csv_encoding=item.csv_encoding,
        csv_delimiter=item.csv_delimiter,
        special_values=item.special_values,
    )


class _SourceReader:
    path: Path
    file_type: str
    headers: tuple[object, ...]
    worksheet: str
    visible_sheets: tuple[str, ...]
    header_row: int
    csv_encoding: str
    csv_delimiter: str

    def iter_data_rows(self, *, max_col: int) -> Iterator[tuple[object, ...]]:
        raise NotImplementedError

    def close(self) -> None:
        raise NotImplementedError


class _ExcelReader(_SourceReader):
    def __init__(self, path: Path, *, worksheet: str, header_row: int):
        self.path = Path(path)
        self.file_type = self.path.suffix.casefold().lstrip(".").upper()
        self.header_row = header_row
        self.csv_encoding = ""
        self.csv_delimiter = ""
        self.workbook = load_workbook(
            self.path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        self.visible_sheets = tuple(
            sheet.title for sheet in self.workbook.worksheets if sheet.sheet_state == "visible"
        )
        if not self.visible_sheets:
            self.workbook.close()
            raise MergeError("没有可见工作表。")
        selected = worksheet or self.visible_sheets[0]
        if selected not in self.visible_sheets:
            self.workbook.close()
            raise MergeError(f"工作表不存在或不可见：{selected}")
        self.worksheet = selected
        self.sheet = self.workbook[selected]
        _reset_read_only_dimensions(self.sheet)
        self.headers = tuple(
            next(
                self.sheet.iter_rows(
                    min_row=self.header_row,
                    max_row=self.header_row,
                    values_only=True,
                ),
                (),
            )
        )

    def iter_data_rows(self, *, max_col: int) -> Iterator[tuple[object, ...]]:
        for row in self.sheet.iter_rows(
            min_row=self.header_row + 1,
            min_col=1,
            max_col=max(max_col, 1),
            values_only=True,
        ):
            yield tuple(row)

    def close(self) -> None:
        self.workbook.close()


class _CsvReader(_SourceReader):
    def __init__(self, path: Path, *, header_row: int, options: CsvOptions):
        self.path = Path(path)
        self.file_type = "CSV"
        self.header_row = header_row
        self.worksheet = ""
        self.visible_sheets = ()
        self.csv_encoding = _detect_csv_encoding(self.path, options.encoding)
        self.csv_delimiter = _detect_csv_delimiter(
            self.path,
            encoding=self.csv_encoding,
            requested=options.delimiter,
        )
        _raise_csv_field_limit()
        self.handle = self.path.open("r", encoding=self.csv_encoding, newline="")
        self.reader = csv.reader(self.handle, delimiter=self.csv_delimiter)
        for _ in range(self.header_row - 1):
            next(self.reader, None)
        self.headers = tuple(next(self.reader, ()))

    def iter_data_rows(self, *, max_col: int) -> Iterator[tuple[object, ...]]:
        del max_col
        for row in self.reader:
            yield tuple(row)

    def close(self) -> None:
        self.handle.close()


def _open_source_reader(
    path: Path,
    *,
    worksheet: str,
    header_row: int,
    csv_options: CsvOptions,
) -> _SourceReader:
    source = Path(path)
    suffix = source.suffix.casefold()
    if suffix in {".xlsx", ".xlsm"}:
        return _ExcelReader(source, worksheet=worksheet, header_row=header_row)
    if suffix == ".csv":
        return _CsvReader(source, header_row=header_row, options=csv_options.validated())
    raise MergeError(f"不支持的输入格式：{source.suffix}")


def _detect_csv_encoding(path: Path, requested: str) -> str:
    if requested != "auto":
        return requested
    sample = path.read_bytes()[:131_072]
    if sample.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    try:
        sample.decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        try:
            sample.decode("gb18030")
            return "gb18030"
        except UnicodeDecodeError as exc:
            raise MergeError("CSV 不是可识别的 UTF-8 或 GB18030 编码。") from exc


def _detect_csv_delimiter(path: Path, *, encoding: str, requested: str) -> str:
    if requested != "auto":
        return requested
    with path.open("r", encoding=encoding, newline="") as handle:
        sample = handle.read(65_536)
    if not sample:
        raise MergeError("CSV 文件为空。")
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error as exc:
        raise MergeError("无法可靠识别 CSV 分隔符，请在文件设置中手动选择。") from exc
    return dialect.delimiter


def _raise_csv_field_limit() -> None:
    limit = 2_147_483_647
    while limit > 131_072:
        try:
            csv.field_size_limit(limit)
            return
        except OverflowError:
            limit //= 10


def _build_unique_header_index(headers: Sequence[object]) -> dict[str, int]:
    index: dict[str, int] = {}
    display: dict[str, str] = {}
    for position, value in enumerate(headers):
        normalized = normalize_header(value)
        if not normalized:
            continue
        if normalized in index:
            shown = str(value).strip() or display[normalized]
            raise MergeError(f"表头存在重复字段，无法确定映射：{shown}")
        index[normalized] = position
        display[normalized] = str(value).strip()
    return index


def _resolve_output_mapping(
    header_index: Mapping[str, int],
    rules: Sequence[OutputColumnRule],
) -> tuple[int | None, ...]:
    mapping: list[int | None] = []
    for rule in rules:
        matches = {
            header_index[normalized]
            for candidate in rule.candidates
            if (normalized := normalize_header(candidate)) in header_index
        }
        if len(matches) > 1:
            raise MergeError(f'输出字段“{rule.name}”同时匹配了多个源字段。')
        if not matches:
            if rule.required:
                raise MergeError(f'缺少必填字段“{rule.name}”。')
            mapping.append(None)
        else:
            mapping.append(next(iter(matches)))
    return tuple(mapping)


def _resolve_special_mapping(
    headers: Sequence[object],
    rules: Sequence[SpecialFieldRule],
    *,
    custom: bool,
) -> tuple[int, ...]:
    if custom:
        header_index = _build_unique_header_index(headers)
    else:
        header_index = {
            normalize_header(value): index
            for index, value in enumerate(headers)
            if normalize_header(value)
        }
    mapping: list[int] = []
    for rule in rules:
        matches = [
            header_index[normalized]
            for candidate in rule.candidates
            if (normalized := normalize_header(candidate)) in header_index
        ]
        unique = set(matches)
        if not unique:
            raise MergeError(f'缺少特殊字段“{rule.name}”。')
        if len(unique) > 1:
            raise MergeError(f'特殊字段“{rule.name}”匹配了多个源字段。')
        if not custom:
            candidate_names = {normalize_header(item) for item in rule.candidates}
            occurrences = [
                index for index, value in enumerate(headers) if normalize_header(value) in candidate_names
            ]
            if len(occurrences) > 1:
                raise MergeError(f'首行存在多个“{rule.name}”字段。')
        mapping.append(next(iter(unique)))
    return tuple(mapping)


def _map_row(row: Sequence[object], mapping: Sequence[int | None]) -> tuple[object, ...]:
    return tuple(
        row[index] if index is not None and index < len(row) else None
        for index in mapping
    )


class _OutputWriter:
    def __init__(
        self,
        *,
        temp_dir: Path,
        headers: Sequence[object],
        column_widths: Sequence[float],
        max_data_rows: int,
    ):
        self.temp_dir = temp_dir
        self.headers = tuple(headers)
        self.column_widths = tuple(column_widths)
        self.max_data_rows = max_data_rows
        self.part_files: list[Path] = []
        self.workbook: Workbook | None = None
        self.worksheet = None
        self.current_data_rows = 0

    def append(self, values: Sequence[object]) -> None:
        if len(values) != len(self.headers):
            raise MergeError("内部错误：输出数据列数与表头不一致。")
        if self.workbook is None or self.current_data_rows >= self.max_data_rows:
            self._close_part()
            self._open_part()
        assert self.worksheet is not None
        self.worksheet.append(self._write_only_row(values))
        self.current_data_rows += 1

    def finish(self) -> list[Path]:
        self._close_part()
        return list(self.part_files)

    def abort(self) -> None:
        if self.worksheet is not None:
            try:
                self.worksheet.close()
            except Exception:
                pass
        if self.workbook is not None:
            try:
                self.workbook.close()
            except Exception:
                pass
        self.workbook = None
        self.worksheet = None

    def _open_part(self) -> None:
        self.workbook = Workbook(write_only=True)
        self.worksheet = self.workbook.create_sheet(OUTPUT_SHEET_NAME)
        self.current_data_rows = 0
        self.worksheet.freeze_panes = "A2"
        self.worksheet.sheet_view.showGridLines = False
        self.worksheet.row_dimensions[1].height = 24
        for index, width in enumerate(self.column_widths, start=1):
            self.worksheet.column_dimensions[get_column_letter(index)].width = width
        header_cells: list[WriteOnlyCell] = []
        for value in self.headers:
            cell = WriteOnlyCell(self.worksheet, value=value)
            cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            header_cells.append(cell)
        self.worksheet.append(header_cells)

    def _close_part(self) -> None:
        if self.workbook is None or self.worksheet is None:
            return
        last_column = get_column_letter(len(self.headers))
        self.worksheet.auto_filter.ref = f"A1:{last_column}{self.current_data_rows + 1}"
        part_path = self.temp_dir / f"part_{len(self.part_files) + 1:04d}.xlsx"
        self.workbook.save(part_path)
        self.workbook.close()
        self.part_files.append(part_path)
        self.workbook = None
        self.worksheet = None

    def _write_only_row(self, values: Sequence[object]) -> list[object]:
        assert self.worksheet is not None
        row: list[object] = []
        for value in values:
            if isinstance(value, str) and value.startswith("="):
                cell = WriteOnlyCell(self.worksheet, value=value)
                cell.data_type = "s"
                row.append(cell)
            else:
                row.append(value)
        return row


def _write_report_workbook(
    path: Path,
    *,
    config: MergeConfig,
    plan: InspectionPlan,
    results: Sequence[FileResult],
    output_paths: Sequence[Path],
    total_rows: int,
    elapsed_seconds: float,
    overwritten_files: Sequence[Path],
) -> None:
    workbook = Workbook()
    workbook.properties.title = f"{APP_NAME}处理报告"
    workbook.properties.creator = f"{APP_NAME} {APP_VERSION}"
    summary = workbook.active
    summary.title = "任务汇总"
    detail = workbook.create_sheet("文件明细")
    mapping = workbook.create_sheet("字段映射")
    successful = sum(1 for item in results if item.status == "成功")
    skipped = sum(1 for item in results if item.status == "跳过")
    excluded = sum(1 for item in results if item.status == "排除")
    summary_rows: list[tuple[object, object]] = [
        ("项目", "内容"),
        ("软件版本", APP_VERSION),
        ("完成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("方案", plan.profile.name),
        ("输入文件夹", str(config.input_dir)),
        ("输出文件夹", str(config.output_dir)),
        ("扫描文件数", len(results)),
        ("成功文件数", successful),
        ("跳过文件数", skipped),
        ("用户排除文件数", excluded),
        ("合并数据行", total_rows),
        ("耗时（秒）", round(elapsed_seconds, 2)),
        ("输出文件", "；".join(item.name for item in output_paths) or "无（没有有效数据）"),
        ("替换旧结果", "；".join(item.name for item in overwritten_files) or "无"),
        ("公式说明", "Excel/XLSM 仅读取上次保存的公式缓存值；未计算公式可能为空。"),
    ]
    for row in summary_rows:
        _append_safe_row(summary, row)
    detail_headers = (
        "状态",
        "相对路径",
        "格式",
        "工作表",
        "表头行",
        "CSV编码",
        "CSV分隔符",
        "特殊字段",
        "写入行数",
        "清理/截断单元格",
        "原因",
    )
    _append_safe_row(detail, detail_headers)
    for result in results:
        specials = "；".join(f"{key}={value}" for key, value in result.special_values)
        delimiter = "制表符" if result.csv_delimiter == "\t" else result.csv_delimiter
        _append_safe_row(
            detail,
            (
                result.status,
                result.relative_path or result.path.name,
                result.file_type,
                result.worksheet,
                result.header_row,
                result.csv_encoding,
                delimiter,
                specials,
                result.rows_written,
                result.adjusted_cells,
                result.reason,
            ),
        )
    mapping_headers = ("类型", "输出/特殊字段", "必填", "可匹配名称或别名", "说明")
    _append_safe_row(mapping, mapping_headers)
    if plan.profile.classic:
        _append_safe_row(mapping, ("经典字段", "A:Q", "是", "按位置及标准表头", "固定保留前17列"))
    else:
        for rule in plan.profile.output_columns:
            _append_safe_row(
                mapping,
                (
                    "输出字段",
                    rule.name,
                    "是" if rule.required else "否",
                    "；".join(rule.candidates),
                    "缺少可选字段时补空",
                ),
            )
    for rule in plan.profile.special_fields if plan.profile.recognize_special_fields else ():
        _append_safe_row(
            mapping,
            ("特殊字段", rule.name, "是", "；".join(rule.candidates), "每个文件必须只有一个唯一非空值"),
        )
    if plan.profile.label.enabled:
        _append_safe_row(
            mapping,
            ("标签列", plan.profile.label.header, "—", plan.profile.label.template, "追加在输出末尾"),
        )
    _style_report_sheet(summary, widths=(18, 80))
    _style_report_sheet(detail, widths=(10, 42, 10, 20, 10, 14, 12, 40, 14, 18, 58))
    _style_report_sheet(mapping, widths=(14, 24, 10, 48, 42))
    for row in detail.iter_rows(min_row=2, max_col=1):
        cell = row[0]
        if cell.value == "成功":
            cell.fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
        elif cell.value == "跳过":
            cell.fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
        else:
            cell.fill = PatternFill(fill_type="solid", fgColor="E7E6E6")
    workbook.save(path)
    workbook.close()


def _append_safe_row(sheet, values: Sequence[object]) -> None:
    sheet.append(list(values))
    row_number = sheet.max_row
    for column, value in enumerate(values, start=1):
        if isinstance(value, str) and value.startswith("="):
            cell = sheet.cell(row=row_number, column=column)
            cell.value = value
            cell.data_type = "s"


def _style_report_sheet(sheet, *, widths: Sequence[float]) -> None:
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 24
    for cell in sheet[1]:
        cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    if sheet.max_column and sheet.max_row:
        sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _reset_read_only_dimensions(sheet) -> None:
    reset = getattr(sheet, "reset_dimensions", None)
    if callable(reset):
        reset()


def _pad_row(row: Sequence[object], width: int) -> tuple[object, ...]:
    values = tuple(row[:width])
    if len(values) < width:
        values += (None,) * (width - len(values))
    return values


def _value_as_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _is_blank_row(row: Sequence[object]) -> bool:
    return all(_is_blank(value) for value in row)


def _prepare_output_values(values: Sequence[object]) -> tuple[tuple[object, ...], int]:
    adjusted = 0
    output: list[object] = []
    for value in values:
        if isinstance(value, str):
            cleaned = _ILLEGAL_XML_CHARS.sub("", value)
            if len(cleaned) > 32_767:
                cleaned = cleaned[:32_767]
            if cleaned != value:
                adjusted += 1
            output.append(cleaned)
        else:
            output.append(value)
    return tuple(output), adjusted


def _display_width(value: object) -> float:
    if value is None:
        return 0.0
    text = _value_as_text(value)
    width = 0.0
    for character in text[:100]:
        width += 2.0 if unicodedata.east_asian_width(character) in {"W", "F"} else 1.0
    return min(width, 40.0)


def _target_output_paths(output_dir: Path, output_name: str, part_count: int) -> list[Path]:
    if part_count == 0:
        return []
    stem = Path(output_name).stem
    if part_count == 1:
        return [output_dir / output_name]
    return [
        output_dir / _PART_NAME_TEMPLATE.format(stem=stem, number=index)
        for index in range(1, part_count + 1)
    ]


def _unique_report_name(output_dir: Path) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = f"{_REPORT_PREFIX}{timestamp}"
    candidate = f"{base}.xlsx"
    counter = 2
    while (output_dir / candidate).exists():
        candidate = f"{base}_{counter}.xlsx"
        counter += 1
    return candidate


def _publish_results(
    *,
    part_files: Sequence[Path],
    target_paths: Sequence[Path],
    report_temp_path: Path,
    report_target_path: Path,
    conflicts: Sequence[Path],
    overwrite: bool,
    temp_dir: Path,
) -> tuple[list[Path], list[Path]]:
    current_conflicts = [path for path in conflicts if path.exists()]
    if current_conflicts and not overwrite:
        raise OutputExistsError(current_conflicts)
    backup_dir = temp_dir / "backup"
    backup_dir.mkdir(exist_ok=True)
    moved_backups: list[tuple[Path, Path]] = []
    published_paths: list[Path] = []
    try:
        for original in current_conflicts:
            backup = backup_dir / f"{uuid.uuid4().hex}_{original.name}"
            original.replace(backup)
            moved_backups.append((original, backup))
        for source, target in zip(part_files, target_paths, strict=True):
            if target.exists():
                if not overwrite:
                    raise OutputExistsError([target])
                backup = backup_dir / f"{uuid.uuid4().hex}_{target.name}"
                target.replace(backup)
                moved_backups.append((target, backup))
            source.replace(target)
            published_paths.append(target)
        report_temp_path.replace(report_target_path)
        published_paths.append(report_target_path)
    except Exception:
        for path in reversed(published_paths):
            if path.exists():
                path.unlink()
        for original, backup in reversed(moved_backups):
            if backup.exists():
                backup.replace(original)
        raise
    return list(target_paths), [original for original, _ in moved_backups]


def _safe_remove_temp_dir(temp_dir: Path, output_dir: Path) -> None:
    try:
        resolved_temp = temp_dir.resolve()
        resolved_output = output_dir.resolve()
        if (
            resolved_temp.parent == resolved_output
            and resolved_temp.name.startswith(".excel_merge_")
            and resolved_temp.is_dir()
        ):
            shutil.rmtree(resolved_temp, ignore_errors=True)
    except OSError:
        pass


def _check_cancelled(cancel_event: threading.Event) -> None:
    if cancel_event.is_set():
        raise MergeCancelled("用户已取消本次任务。")


def _emit(callback: ProgressCallback | None, event: ProgressEvent) -> None:
    if callback is None:
        return
    try:
        callback(event)
    except Exception:
        pass
