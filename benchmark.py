from __future__ import annotations

import argparse
import csv
import ctypes
import os
import threading
import time
from dataclasses import dataclass
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from merge_engine import MergeConfig, merge_workbooks
from merge_profiles import LabelRule, MergeProfile, OutputColumnRule, SpecialFieldRule


HEADERS = ["编号", "姓名", "主讲老师"] + [f"字段{index}" for index in range(4, 18)]


class ProcessMemoryCounters(ctypes.Structure):
    _fields_ = [
        ("cb", ctypes.c_ulong),
        ("PageFaultCount", ctypes.c_ulong),
        ("PeakWorkingSetSize", ctypes.c_size_t),
        ("WorkingSetSize", ctypes.c_size_t),
        ("QuotaPeakPagedPoolUsage", ctypes.c_size_t),
        ("QuotaPagedPoolUsage", ctypes.c_size_t),
        ("QuotaPeakNonPagedPoolUsage", ctypes.c_size_t),
        ("QuotaNonPagedPoolUsage", ctypes.c_size_t),
        ("PagefileUsage", ctypes.c_size_t),
        ("PeakPagefileUsage", ctypes.c_size_t),
    ]


@dataclass(frozen=True)
class BenchmarkResult:
    rows: int
    files: int
    source_format: str
    profile_mode: str
    generation_seconds: float
    merge_seconds: float
    rows_per_second: float
    peak_working_set: int
    output_files: int


def current_working_set() -> int:
    if os.name != "nt":
        return 0
    counters = ProcessMemoryCounters()
    counters.cb = ctypes.sizeof(counters)
    kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
    psapi = ctypes.WinDLL("psapi", use_last_error=True)
    kernel32.GetCurrentProcess.restype = ctypes.c_void_p
    psapi.GetProcessMemoryInfo.argtypes = [
        ctypes.c_void_p,
        ctypes.POINTER(ProcessMemoryCounters),
        ctypes.c_ulong,
    ]
    psapi.GetProcessMemoryInfo.restype = ctypes.c_int
    ok = psapi.GetProcessMemoryInfo(
        kernel32.GetCurrentProcess(),
        ctypes.byref(counters),
        counters.cb,
    )
    return int(counters.WorkingSetSize) if ok else 0


def make_row(identifier: int) -> list[object]:
    return [identifier, f"用户{identifier}", "性能老师"] + [
        identifier * index for index in range(4, 18)
    ]


def create_xlsx(path: Path, start: int, count: int) -> None:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("数据")
    sheet.append(HEADERS)
    for identifier in range(start, start + count):
        sheet.append(make_row(identifier))
    workbook.save(path)
    workbook.close()


def create_csv(path: Path, start: int, count: int) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(HEADERS)
        for identifier in range(start, start + count):
            row = make_row(identifier)
            row[0] = f"{identifier:08d}"
            writer.writerow(row)


def benchmark_profile(mode: str) -> MergeProfile | None:
    if mode == "classic":
        return None
    return MergeProfile(
        name="性能测试自定义方案",
        output_columns=tuple(OutputColumnRule(header) for header in HEADERS),
        special_fields=(SpecialFieldRule("主讲老师"),),
        recognize_special_fields=True,
        label=LabelRule(
            enabled=True,
            header="来源标签",
            template="{文件名不含扩展名}-{特殊字段:主讲老师}",
        ),
    ).validated()


def run_benchmark(
    total_rows: int,
    file_count: int,
    *,
    source_format: str = "xlsx",
    profile_mode: str = "classic",
) -> BenchmarkResult:
    if total_rows < 1 or file_count < 1:
        raise ValueError("行数和文件数必须大于 0。")
    if source_format not in {"xlsx", "csv"}:
        raise ValueError("格式必须是 xlsx 或 csv。")
    if profile_mode not in {"classic", "custom"}:
        raise ValueError("方案必须是 classic 或 custom。")
    if source_format == "csv" and profile_mode == "classic":
        raise ValueError("CSV 性能测试请使用 custom 方案。")

    with TemporaryDirectory(prefix="excel_merge_benchmark_") as temp:
        root = Path(temp)
        input_dir = root / "input"
        output_dir = root / "output"
        input_dir.mkdir()
        base_count, remainder = divmod(total_rows, file_count)
        cursor = 1
        generation_started = time.perf_counter()
        for index in range(file_count):
            count = base_count + (1 if index < remainder else 0)
            suffix = ".xlsx" if source_format == "xlsx" else ".csv"
            path = input_dir / f"测试_{index + 1:02d}{suffix}"
            if source_format == "xlsx":
                create_xlsx(path, cursor, count)
            else:
                create_csv(path, cursor, count)
            cursor += count
        generation_elapsed = time.perf_counter() - generation_started

        stop_sampling = threading.Event()
        peak_working_set = current_working_set()

        def sample_memory() -> None:
            nonlocal peak_working_set
            while not stop_sampling.wait(0.05):
                peak_working_set = max(peak_working_set, current_working_set())

        sampler = threading.Thread(target=sample_memory, daemon=True)
        sampler.start()
        merge_started = time.perf_counter()
        summary = merge_workbooks(
            MergeConfig(
                input_dir,
                output_dir,
                "性能测试结果.xlsx",
                profile=benchmark_profile(profile_mode),
            )
        )
        merge_elapsed = time.perf_counter() - merge_started
        stop_sampling.set()
        sampler.join(timeout=1)

        counted_rows = 0
        expected_columns = 18
        for output in summary.output_files:
            workbook = load_workbook(output, read_only=True, data_only=True)
            sheet = workbook["合并结果"]
            counted_rows += sum(1 for _ in sheet.iter_rows(min_row=2, values_only=True))
            header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            if len(header) != expected_columns:
                raise RuntimeError(f"验收失败：输出列数应为 {expected_columns}，实际 {len(header)}。")
            workbook.close()
        if counted_rows != total_rows:
            raise RuntimeError(f"验收失败：预期 {total_rows:,} 行，实际 {counted_rows:,} 行。")
        return BenchmarkResult(
            rows=total_rows,
            files=file_count,
            source_format=source_format,
            profile_mode=profile_mode,
            generation_seconds=generation_elapsed,
            merge_seconds=merge_elapsed,
            rows_per_second=total_rows / max(merge_elapsed, 0.001),
            peak_working_set=peak_working_set,
            output_files=len(summary.output_files),
        )


def print_result(result: BenchmarkResult) -> None:
    print(f"数据源：{result.source_format.upper()} / {result.profile_mode}")
    print(f"数据规模：{result.rows:,} 行 × 17 列 / {result.files} 个文件")
    print(f"输入生成耗时：{result.generation_seconds:.2f} 秒")
    print(f"合并耗时：{result.merge_seconds:.2f} 秒")
    print(f"处理速度：{result.rows_per_second:,.0f} 行/秒")
    if result.peak_working_set:
        print(f"峰值工作集：{result.peak_working_set / 1024 / 1024:.1f} MiB")
    else:
        print("峰值工作集：当前平台未采集")
    print(f"输出文件数：{result.output_files}")
    print("结果复核：通过")


def main() -> None:
    parser = argparse.ArgumentParser(description="Excel 合并工具 v2 性能验收")
    parser.add_argument("--rows", type=int, default=100_000, help="总数据行数")
    parser.add_argument("--files", type=int, default=2, help="输入文件数")
    parser.add_argument("--format", choices=("xlsx", "csv"), default="xlsx", help="输入格式")
    parser.add_argument(
        "--profile",
        choices=("classic", "custom"),
        default="classic",
        help="合并方案",
    )
    args = parser.parse_args()
    result = run_benchmark(
        args.rows,
        args.files,
        source_format=args.format,
        profile_mode=args.profile,
    )
    print_result(result)


if __name__ == "__main__":
    main()
